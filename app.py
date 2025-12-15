from flask import Flask, render_template_string, request, redirect, url_for, session, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
import os
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import json

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from flask_mysqldb import MySQL
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import os
import json
from functools import wraps
import secrets


# Import utility modules
from utils.email_sender import send_email, send_acceptance_email
from utils.whatsapp_sender import send_whatsapp_notification
from utils.pdf_generator import generate_application_pdf, generate_acceptance_letter
from utils.excel_exporter import export_applications_to_excel


app = Flask(__name__)
app.config['SECRET_KEY'] = 'keem-driving-school-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///keem_driving.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'admin_login'


# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'your_mysql_user'
app.config['MYSQL_PASSWORD'] = 'your_mysql_password'
app.config['MYSQL_DB'] = 'keem_driving_school'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

# Upload Configuration
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {'pdf', 'jpg', 'jpeg', 'png'}

mysql = MySQL(app)

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login to access this page', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


# Database Models
class Admin(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(100), unique=True)
    password = db.Column(db.String(200))
    name = db.Column(db.String(100))
    whatsapp = db.Column(db.String(20))

class Application(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    first_name = db.Column(db.String(100))
    last_name = db.Column(db.String(100))
    email = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    whatsapp = db.Column(db.String(20))
    branch = db.Column(db.String(50))
    language = db.Column(db.String(20))
    license_type = db.Column(db.String(50))
    previous_experience = db.Column(db.String(10))
    medical_conditions = db.Column(db.Text)
    emergency_contact = db.Column(db.String(100))
    emergency_phone = db.Column(db.String(20))
    status = db.Column(db.String(20), default='pending')
    applied_date = db.Column(db.DateTime, default=datetime.utcnow)
    reviewed_date = db.Column(db.DateTime)
    notes = db.Column(db.Text)

class Announcement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200))
    content = db.Column(db.Text)
    created_date = db.Column(db.DateTime, default=datetime.utcnow)
    active = db.Column(db.Boolean, default=True)

@login_manager.user_loader
def load_user(user_id):
    return Admin.query.get(int(user_id))

# Notification Functions
def send_whatsapp_notification(phone, message):
    """Send WhatsApp notification using Twilio API"""
    try:
        # Uncomment and configure when ready to use
        # from twilio.rest import Client
        # account_sid = os.environ.get('TWILIO_ACCOUNT_SID')
        # auth_token = os.environ.get('TWILIO_AUTH_TOKEN')
        # twilio_phone = os.environ.get('TWILIO_PHONE')
        # client = Client(account_sid, auth_token)
        # message = client.messages.create(
        #     from_=f'whatsapp:{twilio_phone}',
        #     to=f'whatsapp:{phone}',
        #     body=message
        # )
        print(f"[WhatsApp] Sent to {phone}: {message}")
        return True
    except Exception as e:
        print(f"[WhatsApp Error]: {e}")
        return False

def send_email_notification(email, subject, body):
    """Send email notification using SendGrid"""
    try:
        # Uncomment and configure when ready to use
        # from sendgrid import SendGridAPIClient
        # from sendgrid.helpers.mail import Mail
        # message = Mail(
        #     from_email='admin@keemdrivingschool.com',
        #     to_emails=email,
        #     subject=subject,
        #     html_content=body
        # )
        # sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
        # response = sg.send(message)
        print(f"[Email] Sent to {email}: {subject}")
        return True
    except Exception as e:
        print(f"[Email Error]: {e}")
        return False

def generate_acceptance_letter_pdf(application):
    """Generate acceptance letter PDF"""
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Header
    c.setFillColorRGB(0.86, 0.15, 0.15)  # Red color
    c.rect(0, height - 100, width, 100, fill=True)
    
    c.setFillColorRGB(1, 1, 1)  # White text
    c.setFont("Helvetica-Bold", 24)
    c.drawString(50, height - 50, "KEEM DRIVING SCHOOL")
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 75, "Excellence in Driver Training")
    
    # Date and Reference
    c.setFillColorRGB(0, 0, 0)  # Black text
    c.setFont("Helvetica", 10)
    c.drawString(50, height - 130, f"Date: {datetime.now().strftime('%B %d, %Y')}")
    c.drawString(50, height - 145, f"Reference: KEEM-{application.id:04d}")
    
    # Applicant Details
    y_position = height - 180
    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y_position, "LETTER OF ACCEPTANCE")
    
    y_position -= 30
    c.setFont("Helvetica", 11)
    c.drawString(50, y_position, f"Dear {application.first_name} {application.last_name},")
    
    y_position -= 30
    c.setFont("Helvetica", 10)
    
    # Acceptance message
    acceptance_text = [
        f"We are pleased to inform you that your application to KEEM Driving School has been ACCEPTED.",
        "",
        f"Branch: {application.branch.upper()}",
        f"License Type: {application.license_type}",
        f"Application ID: KEEM-{application.id:04d}",
        "",
        "Next Steps:",
        "1. Visit our office at your selected branch for registration",
        "2. Bring a valid ID and passport-sized photos (2 copies)",
        "3. Pay the registration fee",
        "4. Receive your training schedule",
        "",
        "Our team will contact you within 2 business days to schedule your orientation session.",
        "",
        "Welcome to KEEM Driving School! We look forward to helping you achieve your driving goals.",
    ]
    
    for line in acceptance_text:
        c.drawString(50, y_position, line)
        y_position -= 20
        if y_position < 100:
            break
    
    # Footer
    y_position = 80
    c.setFont("Helvetica-Bold", 10)
    c.drawString(50, y_position, "Contact Information:")
    c.setFont("Helvetica", 9)
    c.drawString(50, y_position - 15, "Luanshya Branch: Main Street, Luanshya")
    c.drawString(50, y_position - 30, "Mukulushi Branch: Central Road, Mukulushi")
    c.drawString(50, y_position - 45, "Email: info@keemdrivingschool.com | WhatsApp: +260 XXX XXX XXX")
    
    c.save()
    buffer.seek(0)
    return buffer

def generate_applications_excel(applications):
    """Generate Excel file with all applications"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Header styling
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        "ID", "First Name", "Last Name", "Email", "Phone", "WhatsApp",
        "Branch", "License Type", "Experience", "Emergency Contact",
        "Emergency Phone", "Status", "Applied Date", "Language"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for row, app in enumerate(applications, 2):
        ws.cell(row=row, column=1, value=app.id)
        ws.cell(row=row, column=2, value=app.first_name)
        ws.cell(row=row, column=3, value=app.last_name)
        ws.cell(row=row, column=4, value=app.email)
        ws.cell(row=row, column=5, value=app.phone)
        ws.cell(row=row, column=6, value=app.whatsapp)
        ws.cell(row=row, column=7, value=app.branch)
        ws.cell(row=row, column=8, value=app.license_type)
        ws.cell(row=row, column=9, value=app.previous_experience)
        ws.cell(row=row, column=10, value=app.emergency_contact)
        ws.cell(row=row, column=11, value=app.emergency_phone)
        ws.cell(row=row, column=12, value=app.status.upper())
        ws.cell(row=row, column=13, value=app.applied_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=14, value=app.language)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer





# ============== PUBLIC ROUTES ==============

@app.route('/')
def index():
    """Homepage"""
    return render_template('index.html')

@app.route('/about')
def about():
    """About page"""
    return render_template('about.html')

@app.route('/courses')
def courses():
    """Courses page"""
    return render_template('courses.html')

@app.route('/branches')
def branches():
    """Branches information"""
    return render_template('branches.html')

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    """Contact page"""
    if request.method == 'POST':
        try:
            data = request.form.to_dict()
            
            # Insert into database
            cur = mysql.connection.cursor()
            cur.execute("""
                INSERT INTO contact_messages 
                (name, email, phone, subject, message, status, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('name'), data.get('email'), data.get('phone'),
                data.get('subject'), data.get('message'), 'new', datetime.now()
            ))
            
            message_id = cur.lastrowid
            mysql.connection.commit()
            cur.close()
            
            # Send notification to admin
            admin_email = "admin@keemdrivingschool.com"
            email_subject = f"New Contact Message - {data.get('subject')}"
            email_body = f"""
            New contact message received!
            
            From: {data.get('name')}
            Email: {data.get('email')}
            Phone: {data.get('phone')}
            Subject: {data.get('subject')}
            
            Message:
            {data.get('message')}
            
            Message ID: {message_id}
            """
            
            send_email(admin_email, email_subject, email_body)
            
            # Send confirmation to user
            user_message = f"""
            Dear {data.get('name')},
            
            Thank you for contacting KEEM Driving School!
            
            We have received your message and will respond within 24 hours.
            
            Your message:
            {data.get('message')}
            
            Best regards,
            KEEM Driving School Team
            """
            send_email(data.get('email'), "Message Received - KEEM Driving School", user_message)
            
            return jsonify({
                'success': True,
                'message': 'Message sent successfully!'
            })
            
        except Exception as e:
            print(f"Error: {str(e)}")
            return jsonify({
                'success': False,
                'message': 'An error occurred. Please try again.'
            }), 500
    
    return render_template('contact.html')

@app.route('/apply')
def apply():
    """Application form page"""
    return render_template('apply.html')

@app.route('/submit-application', methods=['POST'])
def submit_application():
    """Handle application submission"""
    try:
        # Get form data
        data = request.form.to_dict()
        
        # Handle file upload
        profile_photo = None
        if 'profile_photo' in request.files:
            file = request.files['profile_photo']
            if file and allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                profile_photo = filename
        
        # Insert into database
        cur = mysql.connection.cursor()
        cur.execute("""
            INSERT INTO applications 
            (first_name, last_name, email, phone, whatsapp, date_of_birth, 
             gender, nrc_number, address, city, province, branch, 
             course_type, previous_experience, preferred_language, 
             emergency_contact_name, emergency_contact_phone, 
             medical_conditions, profile_photo, status, created_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data.get('first_name'), data.get('last_name'), data.get('email'),
            data.get('phone'), data.get('whatsapp'), data.get('date_of_birth'),
            data.get('gender'), data.get('nrc_number'), data.get('address'),
            data.get('city'), data.get('province'), data.get('branch'),
            data.get('course_type'), data.get('previous_experience'),
            data.get('preferred_language'), data.get('emergency_contact_name'),
            data.get('emergency_contact_phone'), data.get('medical_conditions'),
            profile_photo, 'pending', datetime.now()
        ))
        
        application_id = cur.lastrowid
        mysql.connection.commit()
        cur.close()
        
        # Send notifications to admin
        admin_email = "admin@keemdrivingschool.com"  # Configure this
        admin_whatsapp = "+260123456789"  # Configure this
        
        email_subject = f"New Application Received - {data.get('first_name')} {data.get('last_name')}"
        email_body = f"""
        New application received!
        
        Name: {data.get('first_name')} {data.get('last_name')}
        Email: {data.get('email')}
        Phone: {data.get('phone')}
        Branch: {data.get('branch')}
        Course: {data.get('course_type')}
        
        Please login to the admin dashboard to review this application.
        Application ID: {application_id}
        """
        
        send_email(admin_email, email_subject, email_body)
        
        whatsapp_message = f"🚗 New application from {data.get('first_name')} {data.get('last_name')} for {data.get('branch')} branch. Application ID: {application_id}"
        send_whatsapp_notification(admin_whatsapp, whatsapp_message)
        
        # Send confirmation to applicant
        applicant_message = f"""
        Dear {data.get('first_name')},
        
        Thank you for applying to KEEM Driving School!
        
        Your application has been received successfully. Application ID: {application_id}
        
        We will review your application and get back to you within 2-3 business days.
        
        Best regards,
        KEEM Driving School Team
        """
        send_email(data.get('email'), "Application Received - KEEM Driving School", applicant_message)
        
        return jsonify({
            'success': True,
            'message': 'Application submitted successfully! Check your email for confirmation.',
            'application_id': application_id
        })
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({
            'success': False,
            'message': 'An error occurred. Please try again.'
        }), 500

@app.route('/application-status/<int:application_id>')
def check_application_status(application_id):
    """Check application status"""
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM applications WHERE id = %s", (application_id,))
    application = cur.fetchone()
    cur.close()
    
    return render_template('application_status.html', application=application)

# ============== ADMIN ROUTES ==============

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login"""
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM admins WHERE email = %s", (email,))
        admin = cur.fetchone()
        cur.close()
        
        if admin and check_password_hash(admin['password'], password):
            session['user_id'] = admin['id']
            session['user_name'] = admin['name']
            session['user_email'] = admin['email']
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid credentials', 'error')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.clear()
    flash('Logged out successfully', 'success')
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    """Admin dashboard"""
    cur = mysql.connection.cursor()
    
    # Get statistics
    cur.execute("SELECT COUNT(*) as total FROM applications")
    total_applications = cur.fetchone()['total']
    
    cur.execute("SELECT COUNT(*) as pending FROM applications WHERE status = 'pending'")
    pending_applications = cur.fetchone()['pending']
    
    cur.execute("SELECT COUNT(*) as accepted FROM applications WHERE status = 'accepted'")
    accepted_applications = cur.fetchone()['accepted']
    
    cur.execute("SELECT COUNT(*) as rejected FROM applications WHERE status = 'rejected'")
    rejected_applications = cur.fetchone()['rejected']
    
    # Get recent applications
    cur.execute("""
        SELECT * FROM applications 
        ORDER BY created_at DESC 
        LIMIT 10
    """)
    recent_applications = cur.fetchall()
    
    cur.close()
    
    stats = {
        'total': total_applications,
        'pending': pending_applications,
        'accepted': accepted_applications,
        'rejected': rejected_applications
    }
    
    return render_template('admin/dashboard.html', 
                         stats=stats, 
                         recent_applications=recent_applications)

@app.route('/admin/applications')
@login_required
def admin_applications():
    """View all applications"""
    status_filter = request.args.get('status', 'all')
    branch_filter = request.args.get('branch', 'all')
    
    cur = mysql.connection.cursor()
    
    query = "SELECT * FROM applications WHERE 1=1"
    params = []
    
    if status_filter != 'all':
        query += " AND status = %s"
        params.append(status_filter)
    
    if branch_filter != 'all':
        query += " AND branch = %s"
        params.append(branch_filter)
    
    query += " ORDER BY created_at DESC"
    
    cur.execute(query, tuple(params))
    applications = cur.fetchall()
    cur.close()
    
    return render_template('admin/applications.html', 
                         applications=applications,
                         status_filter=status_filter,
                         branch_filter=branch_filter)

@app.route('/admin/application/<int:application_id>')
@login_required
def view_application(application_id):
    """View single application details"""
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM applications WHERE id = %s", (application_id,))
    application = cur.fetchone()
    cur.close()
    
    if not application:
        flash('Application not found', 'error')
        return redirect(url_for('admin_applications'))
    
    return render_template('admin/view_application.html', application=application)

@app.route('/admin/application/<int:application_id>/update-status', methods=['POST'])
@login_required
def update_application_status(application_id):
    """Update application status"""
    try:
        status = request.form.get('status')
        notes = request.form.get('notes', '')
        
        cur = mysql.connection.cursor()
        cur.execute("""
            UPDATE applications 
            SET status = %s, admin_notes = %s, updated_at = %s
            WHERE id = %s
        """, (status, notes, datetime.now(), application_id))
        
        # Get application details
        cur.execute("SELECT * FROM applications WHERE id = %s", (application_id,))
        application = cur.fetchone()
        
        mysql.connection.commit()
        cur.close()
        
        # If accepted, send acceptance letter
        if status == 'accepted' and application:
            # Generate acceptance letter PDF
            pdf_path = generate_acceptance_letter(application)
            
            # Send email with acceptance letter
            subject = "Congratulations! Your Application Has Been Accepted - KEEM Driving School"
            send_acceptance_email(application['email'], subject, application, pdf_path)
            
            # Send WhatsApp notification
            whatsapp_msg = f"""
🎉 Congratulations {application['first_name']}!

Your application to KEEM Driving School has been ACCEPTED!

Please check your email for the acceptance letter with further details.

Welcome to KEEM Driving School!
            """
            send_whatsapp_notification(application['whatsapp'], whatsapp_msg)
        
        flash('Application status updated successfully', 'success')
        return redirect(url_for('view_application', application_id=application_id))
        
    except Exception as e:
        print(f"Error: {str(e)}")
        flash('An error occurred', 'error')
        return redirect(url_for('view_application', application_id=application_id))

@app.route('/admin/export/pdf/<int:application_id>')
@login_required
def export_application_pdf(application_id):
    """Export single application as PDF"""
    cur = mysql.connection.cursor()
    cur.execute("SELECT * FROM applications WHERE id = %s", (application_id,))
    application = cur.fetchone()
    cur.close()
    
    if not application:
        flash('Application not found', 'error')
        return redirect(url_for('admin_applications'))
    
    pdf_path = generate_application_pdf(application)
    return send_file(pdf_path, as_attachment=True)

@app.route('/admin/export/excel')
@login_required
def export_applications_excel():
    """Export all applications as Excel"""
    status = request.args.get('status', 'all')
    branch = request.args.get('branch', 'all')
    
    cur = mysql.connection.cursor()
    
    query = "SELECT * FROM applications WHERE 1=1"
    params = []
    
    if status != 'all':
        query += " AND status = %s"
        params.append(status)
    
    if branch != 'all':
        query += " AND branch = %s"
        params.append(branch)
    
    cur.execute(query, tuple(params))
    applications = cur.fetchall()
    cur.close()
    
    excel_path = export_applications_to_excel(applications)
    return send_file(excel_path, as_attachment=True)

@app.route('/admin/settings')
@login_required
def admin_settings():
    """Admin settings"""
    return render_template('admin/settings.html')

# ============== API ROUTES ==============

@app.route('/api/stats')
@login_required
def api_stats():
    """Get dashboard statistics"""
    cur = mysql.connection.cursor()
    
    cur.execute("SELECT COUNT(*) as count FROM applications WHERE status = 'pending'")
    pending = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM applications WHERE status = 'accepted'")
    accepted = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM applications WHERE status = 'rejected'")
    rejected = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM applications WHERE created_at >= %s", 
                (datetime.now() - timedelta(days=7),))
    this_week = cur.fetchone()['count']
    
    cur.close()
    
    return jsonify({
        'pending': pending,
        'accepted': accepted,
        'rejected': rejected,
        'this_week': this_week
    })

# ============== ERROR HANDLERS ==============

@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(e):
    return render_template('errors/500.html'), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)