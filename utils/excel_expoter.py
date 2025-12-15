"""
Excel Export Utility
Install: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

def export_applications_to_excel(applications):
    """
    Export applications to Excel file
    """
    # Create exports directory
    os.makedirs('exports/excel', exist_ok=True)
    
    filename = f"exports/excel/applications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    
    # Define headers
    headers = [
        'ID', 'First Name', 'Last Name', 'Email', 'Phone', 'WhatsApp',
        'Date of Birth', 'Gender', 'NRC Number', 'Address', 'City', 'Province',
        'Branch', 'Course Type', 'Preferred Language', 'Previous Experience',
        'Emergency Contact', 'Emergency Phone', 'Medical Conditions',
        'Status', 'Date Applied', 'Admin Notes'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = cell_border
    
    # Write data
    for row_num, app in enumerate(applications, 2):
        data = [
            app.get('id', ''),
            app.get('first_name', ''),
            app.get('last_name', ''),
            app.get('email', ''),
            app.get('phone', ''),
            app.get('whatsapp', ''),
            str(app.get('date_of_birth', '')),
            app.get('gender', ''),
            app.get('nrc_number', ''),
            app.get('address', ''),
            app.get('city', ''),
            app.get('province', ''),
            app.get('branch', ''),
            app.get('course_type', ''),
            app.get('preferred_language', ''),
            app.get('previous_experience', ''),
            app.get('emergency_contact_name', ''),
            app.get('emergency_contact_phone', ''),
            app.get('medical_conditions', ''),
            app.get('status', ''),
            str(app.get('created_at', '')),
            app.get('admin_notes', '')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = cell_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            # Color code by status
            if col_num == 20:  # Status column
                if value == 'accepted':
                    cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                elif value == 'rejected':
                    cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                elif value == 'pending':
                    cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Adjust column widths
    column_widths = {
        'A': 8, 'B': 15, 'C': 15, 'D': 25, 'E': 15, 'F': 15,
        'G': 12, 'H': 10, 'I': 15, 'J': 30, 'K': 15, 'L': 15,
        'M': 12, 'N': 20, 'O': 12, 'P': 25, 'Q': 20, 'R': 15,
        'S': 25, 'T': 12, 'U': 18, 'V': 30
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(filename)
    return filename

def export_students_to_excel(students):
    """
    Export students to Excel file
    """
    os.makedirs('exports/excel', exist_ok=True)
    
    filename = f"exports/excel/students_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"
    
    headers = [
        'ID', 'Student Number', 'Full Name', 'Email', 'Phone',
        'Branch', 'Course', 'Enrollment Date', 'Start Date', 'End Date',
        'Status', 'Payment Status', 'Total Fee', 'Amount Paid', 'Balance'
    ]
    
    # Apply header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write student data
    for row_num, student in enumerate(students, 2):
        balance = float(student.get('total_fee', 0)) - float(student.get('amount_paid', 0))
        
        data = [
            student.get('id', ''),
            student.get('student_number', ''),
            f"{student.get('first_name', '')} {student.get('last_name', '')}",
            student.get('email', ''),
            student.get('phone', ''),
            student.get('branch', ''),
            student.get('course_type', ''),
            str(student.get('enrollment_date', '')),
            str(student.get('course_start_date', '')),
            str(student.get('course_end_date', '')),
            student.get('status', ''),
            student.get('payment_status', ''),
            student.get('total_fee', ''),
            student.get('amount_paid', ''),
            balance
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(filename)
    return filename

def export_payments_to_excel(payments):
    """
    Export payment records to Excel
    """
    os.makedirs('exports/excel', exist_ok=True)
    
    filename = f"exports/excel/payments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    headers = [
        'Payment ID', 'Student Number', 'Student Name', 'Amount',
        'Payment Method', 'Reference', 'Payment Date', 'Received By', 'Notes'
    ]
    
    # Apply styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_num, payment in enumerate(payments, 2):
        data = [
            payment.get('id', ''),
            payment.get('student_number', ''),
            payment.get('student_name', ''),
            payment.get('amount', ''),
            payment.get('payment_method', ''),
            payment.get('payment_reference', ''),
            str(payment.get('payment_date', '')),
            payment.get('received_by', ''),
            payment.get('notes', '')
        ]
        
        for col_num, value in enumerate(data, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.freeze_panes = 'A2'
    wb.save(filename)
    return filename